"""Exportación limpia a Excel (openpyxl)."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1A1C22", end_color="1A1C22", fill_type="solid")
HEADER_FONT = Font(color="7FD8BE", bold=True, name="Calibri")
BODY_FONT = Font(color="1A1A1A", name="Calibri")


def export_results(
    results: list[dict],
    output_path: str | Path,
    domain_keys: list[str],
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    base_headers = [
        "source_file",
        "texto_original",
        "texto_optimizado",
        "tokens_original",
        "tokens_optimizado",
        "ahorro_tokens",
        "ahorro_pct",
        "optimizado",
        "metodo_clasificacion",
    ]
    headers = base_headers + list(domain_keys)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(results, 2):
        cls: dict[str, Any] = r.get("classification") or {}
        row_vals = [
            r.get("source", ""),
            r.get("text", ""),
            r.get("text_optimized", ""),
            r.get("tokens_original", 0),
            r.get("tokens_optimized", 0),
            r.get("savings_tokens", 0),
            r.get("savings_pct", 0),
            "si" if r.get("optimized") else "no",
            cls.get("_method", r.get("method", "rules")),
        ]
        for key in domain_keys:
            val = cls.get(key, "")
            if val is None:
                val = ""
            row_vals.append(val)

        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = BODY_FONT

    # Anchos razonables
    widths = {
        "A": 18, "B": 48, "C": 48, "D": 14, "E": 16, "F": 14, "G": 12, "H": 12, "I": 18,
    }
    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(letter, 16)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    wb.save(path)
    return path


def export_summary_sheet(wb: Workbook, summary: dict) -> None:
    ws = wb.create_sheet("Resumen")
    rows = [
        ("total_items", summary.get("total_items")),
        ("tokens_original", summary.get("tokens_original")),
        ("tokens_optimized", summary.get("tokens_optimized")),
        ("savings_tokens", summary.get("savings_tokens")),
        ("cost_original_usd", summary.get("cost_original_usd")),
        ("cost_optimized_usd", summary.get("cost_optimized_usd")),
        ("cost_savings_usd", summary.get("cost_savings_usd")),
        ("elapsed_ms", summary.get("elapsed_ms")),
    ]
    ws.cell(row=1, column=1, value="metrica").font = HEADER_FONT
    ws.cell(row=1, column=2, value="valor").font = HEADER_FONT
    for i, (k, v) in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)

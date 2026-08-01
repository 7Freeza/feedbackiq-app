"""Ingesta robusta de Excel: streaming read_only, límite de filas, filtro basura."""

from __future__ import annotations

import re
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import pandas as pd

from app.config import get_settings

TEXT_KEYWORDS = [
    "review", "reseña", "resena", "comentario", "feedback", "texto", "opinion",
    "text", "comment", "incidencia", "descripcion", "descripción", "detalle",
    "mensaje", "nota", "observacion", "observación", "mensaje_texto",
    "solicitud", "clausula", "cláusula", "contenido", "body", "body_text",
    "content", "narrativa", "descripcion_problema", "problema",
]

_STOPWORDS = {
    "de", "la", "el", "los", "las", "un", "una", "unos", "unas", "y", "o", "a",
    "al", "del", "con", "por", "para", "en", "que", "mi", "tu", "su", "no", "se",
}


@dataclass
class IngestFileResult:
    filename: str
    texts: list[str] = field(default_factory=list)
    column_detected: str | None = None
    row_count: int = 0
    error: str | None = None
    truncated: bool = False


@dataclass
class IngestResult:
    items: list[tuple[str, str]]  # (source_file, text)
    files: list[IngestFileResult]
    total_rows: int
    errors: list[str]


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.lower().strip()


def _clean_cell(v) -> str:
    return re.sub(r"\s+", " ", str(v).strip())


def _looks_like_header_list(text: str) -> bool:
    if any(c in text for c in [",", ".", ";", ":", "!", "?"]):
        return False
    words = text.split()
    if len(words) < 2:
        return False
    if any(len(w) > 12 for w in words):
        return False
    content = [w for w in words if w.lower() not in _STOPWORDS]
    if not content or len(content) < len(words):
        return False
    return True


def _is_junk(text: str) -> bool:
    if len(text) < 3:
        return True
    if sum(1 for c in text if c.isalpha()) == 0:
        return True
    if re.fullmatch(r"[\d/:\-.\s]+", text):
        return True
    if len(text) < 10:
        return True
    if _looks_like_header_list(text):
        return True
    return False


def _column_quality(header, values) -> tuple[float, list[str]]:
    norm = _norm(header)
    kw_hits = sum(1 for kw in TEXT_KEYWORDS if kw in norm)
    cleaned = [
        _clean_cell(v)
        for v in values
        if v is not None and not _is_junk(_clean_cell(v))
    ]
    avg_len = sum(len(c) for c in cleaned) / max(len(cleaned), 1)
    return kw_hits * 100 + avg_len, cleaned


def _best_column_from_df(df: pd.DataFrame) -> tuple[list[str], str | None]:
    best_quality = -1.0
    best: list[str] = []
    best_col: str | None = None
    for col in df.columns:
        values = df[col].dropna().astype(str).tolist()
        quality, cleaned = _column_quality(str(col), values)
        if quality > best_quality and cleaned:
            best_quality = quality
            best = cleaned
            best_col = str(col)
    return best, best_col


def _extract_streaming(path: Path, max_rows: int) -> IngestFileResult:
    """Lectura openpyxl read_only + iter_rows (menor huella de memoria)."""
    result = IngestFileResult(filename=path.name)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            result.error = f"Hoja vacía: {path.name}"
            return result

        headers = [
            str(h).strip() if h is not None else f"col_{i}"
            for i, h in enumerate(header_row)
        ]
        # Acumular valores por columna hasta max_rows * 1.2 para scoring
        cols: list[list] = [[] for _ in headers]
        raw_count = 0
        for row in rows_iter:
            raw_count += 1
            for ci in range(len(headers)):
                if ci < len(row):
                    cols[ci].append(row[ci])
            if raw_count >= max_rows * 2:
                # suficiente muestra para detectar columna
                break

        best_q = -1.0
        best_texts: list[str] = []
        best_col = None
        for ci, h in enumerate(headers):
            q, cleaned = _column_quality(h, cols[ci])
            if q > best_q and cleaned:
                best_q = q
                best_texts = cleaned
                best_col = h

        if not best_texts:
            result.error = (
                f"No se detectó columna de texto en {path.name}. "
                "Renombra la columna (review, reseña, comentario, texto…)."
            )
            return result

        # Segunda pasada solo de la mejor columna si se truncó la muestra
        # (si ya tenemos max_rows, usamos lo filtrado)
        texts = [t for t in best_texts if len(t) >= 10][:max_rows]
        truncated = len(best_texts) > max_rows or raw_count >= max_rows * 2

        # Si la primera pasada se cortó y tenemos pocos textos, re-leer solo col ganadora
        if len(texts) < min(100, max_rows) and best_col is not None:
            ci = headers.index(best_col) if best_col in headers else 0
            texts = []
            wb2 = openpyxl.load_workbook(path, data_only=True, read_only=True)
            try:
                ws2 = wb2.active
                it = ws2.iter_rows(values_only=True)
                next(it, None)  # skip header
                for row in it:
                    if ci < len(row) and row[ci] is not None:
                        cell = _clean_cell(row[ci])
                        if not _is_junk(cell) and len(cell) >= 10:
                            texts.append(cell)
                            if len(texts) >= max_rows:
                                truncated = True
                                break
            finally:
                wb2.close()

        result.texts = texts
        result.column_detected = best_col
        result.row_count = len(texts)
        result.truncated = truncated
        if not result.texts:
            result.error = f"Filas de texto válidas insuficientes en {path.name}"
        return result
    finally:
        wb.close()


def extract_from_excel(file_path: str | Path, max_rows: int | None = None) -> IngestFileResult:
    path = Path(file_path)
    result = IngestFileResult(filename=path.name)
    settings = get_settings()
    limit = max_rows if max_rows is not None else settings.MAX_ROWS

    if not path.exists():
        result.error = f"Archivo no encontrado: {path.name}"
        return result
    if path.suffix.lower() not in {".xlsx", ".xls"}:
        result.error = f"Formato no soportado: {path.suffix}"
        return result
    if path.stat().st_size == 0:
        result.error = f"Archivo vacío: {path.name}"
        return result

    size_mb = path.stat().st_size / (1024 * 1024)
    # Archivos medianos/grandes: streaming; chicos: pandas (más simple)
    use_stream = size_mb >= 1.5 or path.stat().st_size > 800_000

    try:
        if use_stream:
            return _extract_streaming(path, limit)

        df = pd.read_excel(path, dtype=str, engine="openpyxl")
        if df.empty:
            result.error = f"Excel sin filas de datos: {path.name}"
            return result
        texts, col = _best_column_from_df(df)
        del df
        if not texts:
            df2 = pd.read_excel(path, dtype=str, header=None, engine="openpyxl")
            df2.columns = [f"col_{i}" for i in range(df2.shape[1])]
            texts, col = _best_column_from_df(df2)
            del df2
        if not texts:
            return _extract_streaming(path, limit)

        cleaned = [t for t in texts if len(t) >= 10]
        truncated = len(cleaned) > limit
        result.texts = cleaned[:limit]
        result.column_detected = col
        result.row_count = len(result.texts)
        result.truncated = truncated
        if not result.texts:
            result.error = f"Filas de texto válidas insuficientes en {path.name}"
        return result
    except Exception as exc:
        result.error = f"Excel corrupto o ilegible ({path.name}): {exc}"
        return result


def extract_from_files(paths: list[str | Path], max_workers: int = 4) -> IngestResult:
    """Procesa uno o varios .xlsx. Un archivo malo no tumba el lote."""
    settings = get_settings()
    files: list[IngestFileResult] = []
    items: list[tuple[str, str]] = []
    errors: list[str] = []
    remaining = settings.MAX_ROWS

    path_list = list(paths)
    if not path_list:
        return IngestResult(items=[], files=[], total_rows=0, errors=["No se recibieron archivos"])

    workers = min(max_workers, max(1, len(path_list)))
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {
            pool.submit(extract_from_excel, p, settings.MAX_ROWS): p for p in path_list
        }
        for fut in as_completed(futures):
            fr = fut.result()
            files.append(fr)
            if fr.error:
                errors.append(fr.error)

    files.sort(key=lambda f: f.filename)

    for fr in files:
        if fr.error:
            continue
        for t in fr.texts:
            if remaining <= 0:
                fr.truncated = True
                errors.append(
                    f"Límite de {settings.MAX_ROWS} filas alcanzado; "
                    f"resto de {fr.filename} omitido."
                )
                break
            items.append((fr.filename, t))
            remaining -= 1

    return IngestResult(
        items=items,
        files=files,
        total_rows=len(items),
        errors=errors,
    )


def validate_upload_meta(filename: str, size: int, max_bytes: int) -> str | None:
    if not filename.lower().endswith((".xlsx", ".xls")):
        return f"Tipo de archivo no permitido: {filename}. Solo .xlsx / .xls"
    if size <= 0:
        return f"Archivo vacío: {filename}"
    if size > max_bytes:
        mb = max_bytes // (1024 * 1024)
        return f"{filename} excede el tamaño máximo ({mb} MB)"
    return None
